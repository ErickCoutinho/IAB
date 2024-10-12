from openpyxl import load_workbook


def ler_coluna_conveniado(excel_file, aba):
    """
    Lê a coluna 'Conveniado' de uma planilha Excel, independentemente de qual coluna ela esteja.

    :param excel_file: Caminho para o arquivo Excel.
    :param aba: Nome da aba a ser lida.
    :return: Uma lista contendo os valores da coluna 'Conveniado' ou None se a coluna não for encontrada.
    """
    # Carregar o arquivo Excel com openpyxl
    workbook = load_workbook(excel_file)
    worksheet = workbook[aba]
    # Localizar a coluna que contém 'Conveniado' na primeira linha (cabeçalho)
    col_conveniado = None
    for col in worksheet.iter_cols(1, worksheet.max_column):
        if col[0].value and 'Conveniado' in col[0].value:
            col_conveniado = col[0].column_letter
            break
    if not col_conveniado:
        print("Erro: A coluna 'Conveniado' não foi encontrada.")
        return None
    # Extrair os valores da coluna 'Conveniado'
    conveniados = []
    for row in range(2, worksheet.max_row + 1):
        conveniado = worksheet[f'{col_conveniado}{row}'].value
        if conveniado:
            conveniados.append(conveniado.strip())

    return conveniados
