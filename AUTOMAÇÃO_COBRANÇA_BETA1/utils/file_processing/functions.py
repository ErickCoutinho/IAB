import pandas as pd
from tkinter import filedialog, messagebox
from src.final_returns.returns_names import (filtrar_nomes_finais, filtrar_nomes_atrasados, filtrar_nomes_cartorio,
                                             filtrar_nomes_devolvido, filtrar_nomes_tarifas,
                                             filtrar_nomes_nao_encontrados)
from openpyxl import load_workbook
from interface.loading import fechar_loading
from utils.extract.extract_excel import ler_nomes_excel


def carregar_arquivo_txt(excel_file=None, aba_selecionada=None):
    file_path = filedialog.askopenfilename(title="Selecione o arquivo .txt", filetypes=[("Arquivo TXT", "*.txt")])
    if file_path:
        try:
            # Chama a função filtrar_nomes_finais para processar o arquivo TXT e obter o dicionário filtrado
            dicionario_nomes_valores = filtrar_nomes_finais(file_path)
            dicionario_atrasados = filtrar_nomes_atrasados(file_path)
            dicionario_cartorio = filtrar_nomes_cartorio(file_path)
            dicionario_devolvido = filtrar_nomes_devolvido(file_path)
            dicionario_tarifas = filtrar_nomes_tarifas(file_path)
            dicionario_nao_encontrados = filtrar_nomes_nao_encontrados(file_path,
                                                                       ler_nomes_excel(excel_file, aba_selecionada))
            print("Arquivo TXT carregado e processado com sucesso.")
            return (dicionario_nomes_valores, dicionario_atrasados, dicionario_cartorio, dicionario_devolvido,
                    dicionario_tarifas, dicionario_nao_encontrados)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar o arquivo TXT: {str(e)}")
        return None


# Função para carregar o arquivo Excel e escolher a aba
def carregar_arquivo_excel(combobox_aba):
    file_path = filedialog.askopenfilename(title="Selecione o arquivo Excel", filetypes=[("Arquivo Excel", "*.xlsx")])
    if file_path:
        try:
            excel_file = pd.ExcelFile(file_path)
            # Populando o combobox com os nomes das abas
            combobox_aba['values'] = excel_file.sheet_names
            combobox_aba.current(0)  # Seleciona a primeira aba por padrão
            print("Arquivo Excel carregado com sucesso.")
            return excel_file
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar o arquivo Excel: {str(e)}")
    return None


def processar_dados(excel_file, aba_selecionada, dicionario_nomes_valores):
    """
    Função PRINCIPAL para processar os dados no arquivo Excel, comparando os nomes na coluna "Conveniado"
    com os nomes de um dicionário e atualizando a coluna "Total fatura titular" com os valores correspondentes.

    Parâmetros:
    - excel_file: Caminho do arquivo Excel a ser processado.
    - aba_selecionada: Nome da aba no Excel onde o processamento será feito.
    - dicionario_nomes_valores: Dicionário contendo os nomes e os valores a serem comparados e inseridos no Excel.

    Fluxo da função:
    1. Carregar o arquivo Excel e a aba especificada mantendo a formatação existente.
    2. Identificar as colunas "Conveniado" e "Total fatura titular" dinamicamente.
    3. Comparar os nomes da coluna "Conveniado" com os nomes do dicionário.
    4. Atualizar a coluna "Total fatura titular" com os valores do dicionário quando houver correspondência.
    5. Exibir os nomes encontrados e atualizados.
    6. Salvar o arquivo Excel atualizado, sobrescrevendo o original.
    """
    dicionario_nomes_correspondetes = {}
    try:
        # Carregar o workbook existente
        workbook = load_workbook(excel_file)
        worksheet = workbook[aba_selecionada]

        # Identificar colunas "Conveniado" e "Total fatura titular"
        col_conveniado = None
        col_total_fatura = None
        for col in worksheet.iter_cols(1, worksheet.max_column):
            header_value = col[0].value
            if header_value:
                if 'Conveniado' in header_value:
                    col_conveniado = col[0].column_letter
                elif 'Total fatura titular' in header_value:
                    col_total_fatura = col[0].column_letter
            if col_conveniado and col_total_fatura:
                break

        # Verificar se as colunas foram encontradas
        if not col_conveniado:
            messagebox.showerror("Erro", "A coluna 'Conveniado' não foi encontrada.")
            return
        if not col_total_fatura:
            messagebox.showerror("Erro", "A coluna 'Total fatura titular' não foi encontrada.")
            return

        # Processar os dados e atualizar a planilha
        nomes_encontrados = []
        nomes_com_valor_zero = []  # Lista para armazenar nomes com valor igual a zero
        for nome, valor in dicionario_nomes_valores.items():
            for row in range(2, worksheet.max_row + 1):
                conveniado = worksheet[f'{col_conveniado}{row}'].value
                if conveniado and nome.upper() in conveniado.upper():
                    worksheet[f'{col_total_fatura}{row}'] = valor
                    nomes_encontrados.append(nome)
                    dicionario_nomes_correspondetes[nome] = valor
                    if valor == 0:
                        nomes_com_valor_zero.append(nome)

        # Alerta se houver nomes com valor igual a zero
        if nomes_com_valor_zero:
            messagebox.showinfo("ATENÇÃO",
                                f"Verificar seguintes nomes no TXT: {', '.join(nomes_com_valor_zero)}")

        # Sobrescrever o arquivo Excel original
        workbook.save(excel_file)
        messagebox.showinfo("Arquivo Salvo", f"Arquivo sobrescrito em {excel_file}")

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao processar os dados: {str(e)}")
    finally:
        fechar_loading()

    return dicionario_nomes_correspondetes
