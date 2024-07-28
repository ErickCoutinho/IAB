import re
import unicodedata
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy
import tkinter as tk

def normalize_name(name):
    try:
        name = unicodedata.normalize('NFKD', name)
        name = ''.join([c for c in name if not unicodedata.combining(c)])
        return name.strip().upper()
    except Exception as e:
        return f"Erro ao normalizar nome: {str(e)}"

def extrair_nomes_debito_automatico(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.readlines()

        nomes = []
        for line in content[1:]:
            try:
                nome = line[1:40].strip().upper()
                nome_limpo = normalize_name(re.sub(r'\d', '', nome).strip())
                nomes.append(nome_limpo)
                print(f"Nome extraído: {nome_limpo}")
            except Exception as e:
                print(f"Erro ao extrair nome: {str(e)}")
                continue

        return nomes
    except Exception as e:
        print(f"Erro ao extrair nomes de débito automático: {str(e)}")
        return []

def processar_nomes(ws, nomes, pago_col_index, dia_pgto_col_index, taxa_col_index, data_posicao, taxas, nomes_debito_automatico):
    try:
        total_fatura_titular_col_index = next((i for i, cell in enumerate(ws[1], 1) if cell.value == "Total fatura titular"), None)
        nomes_encontrados = []
        valores_associados = []

        for nome_procurado in nomes:
            try:
                encontrado = False
                nome_procurado_limpo = normalize_name(nome_procurado.strip())
                print(f"Procurando por: {nome_procurado_limpo}")
                for row in ws.iter_rows(min_row=2, values_only=False):
                    try:
                        nome_excel = normalize_name(str(row[0].value).strip())
                        if nome_procurado_limpo == nome_excel or nome_procurado_limpo in nomes_debito_automatico:
                            nomes_encontrados.append(nome_procurado)
                            valor = row[total_fatura_titular_col_index - 1].value
                            valor_formatado = f"{valor:.2f}" if isinstance(valor, (int, float)) else str(valor) if valor else "Valor não encontrado"
                            valores_associados.append(valor_formatado)
                            row[pago_col_index - 1].value = "X"
                            row[dia_pgto_col_index - 1].value = data_posicao
                            row[taxa_col_index - 1].value = f"{taxas.get(nome_procurado, 0):.2f}"
                            print(f"Encontrado: {nome_procurado_limpo} com valor {valor_formatado}")
                            encontrado = True
                            break
                    except Exception as e:
                        print(f"Erro ao processar linha do Excel: {str(e)}")
                        continue
                if not encontrado:
                    print(f"Não encontrado: {nome_procurado_limpo}")
                    valores_associados.append("Valor não encontrado")
            except Exception as e:
                print(f"Erro ao processar nome: {str(e)}")
                continue
        return nomes_encontrados, valores_associados
    except Exception as e:
        print(f"Erro ao processar nomes no Excel: {str(e)}")
        return [], []

def buscar_nomes_excel(file_path_excel, nomes, aba, data_posicao, taxas, nomes_debito_automatico):
    try:
        wb = load_workbook(filename=file_path_excel, data_only=False)
        ws = wb[aba]

        pago_col_label = "Pago?"
        pago_col_index = None
        for col in range(1, ws.max_column + 2):
            if col <= ws.max_column and ws.cell(row=1, column=col).value == pago_col_label:
                pago_col_index = col
                break
        if not pago_col_index:
            pago_col_index = ws.max_column + 1
            ws.cell(row=1, column=pago_col_index).value = pago_col_label

        dia_pgto_col_label = "DIA PGTO"
        dia_pgto_col_index = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == dia_pgto_col_label:
                dia_pgto_col_index = col
                break
        if not dia_pgto_col_index:
            dia_pgto_col_index = ws.max_column + 1
            ws.cell(row=1, column=dia_pgto_col_index).value = dia_pgto_col_label

        taxa_col_label = "Taxa"
        taxa_col_index = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == taxa_col_label:
                taxa_col_index = col
                break
        if not taxa_col_index:
            taxa_col_index = ws.max_column + 1
            ws.cell(row=1, column=taxa_col_index).value = taxa_col_label

        source_col_index = pago_col_index - 1
        copiar_formatacao_e_mesclagens(ws, source_col_index, pago_col_index)
        copiar_formatacao_e_mesclagens(ws, source_col_index, dia_pgto_col_index)
        copiar_formatacao_e_mesclagens(ws, source_col_index, taxa_col_index)

        nomes_encontrados, valores_associados = processar_nomes(ws, nomes, pago_col_index, dia_pgto_col_index, taxa_col_index, data_posicao, taxas, nomes_debito_automatico)
        wb.save(filename=file_path_excel)
        wb.close()
        return nomes_encontrados, valores_associados
    except Exception as e:
        print(f"Erro ao buscar nomes no Excel: {str(e)}")
        return [], []

def copiar_formatacao_e_mesclagens(ws, source_col_index, target_col_index):
    try:
        for row in range(1, ws.max_row + 1):
            try:
                source_cell = ws.cell(row=row, column=source_col_index)
                target_cell = ws.cell(row=row, column=target_col_index)
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.number_format = copy(source_cell.number_format)
                    target_cell.protection = copy(source_cell.protection)
                    target_cell.alignment = copy(source_cell.alignment)
            except Exception as e:
                print(f"Erro ao copiar formatação: {str(e)}")
                continue

        new_merges = []
        for merge_cell in ws.merged_cells.ranges:
            if merge_cell.min_col <= source_col_index <= merge_cell.max_col:
                new_range = f"{get_column_letter(target_col_index)}{merge_cell.min_row}:{get_column_letter(target_col_index)}{merge_cell.max_row}"
                new_merges.append(new_range)

        for merge_range in new_merges:
            ws.merge_cells(merge_range)
    except Exception as e:
        print(f"Erro ao copiar mesclagens: {str(e)}")

def processar_arquivo_ret(file_path_ret, file_path_excel, aba, txt_area):
    try:
        nomes_debito_automatico = extrair_nomes_debito_automatico(file_path_ret)
        nomes_debito_automatico = [normalize_name(nome) for nome in nomes_debito_automatico]

        wb = load_workbook(filename=file_path_excel, data_only=False)
        ws = wb[aba]

        pago_col_label = "Pago?"
        pago_col_index = None
        for col in range(1, ws.max_column + 2):
            if col <= ws.max_column and ws.cell(row=1, column=col).value == pago_col_label:
                pago_col_index = col
                break
        if not pago_col_index:
            pago_col_index = ws.max_column + 1
            ws.cell(row=1, column=pago_col_index).value = pago_col_label

        resultados = []

        for nome_procurado in nomes_debito_automatico:
            try:
                encontrado = False
                nome_procurado_limpo = normalize_name(nome_procurado.strip())
                print(f"Procurando por: {nome_procurado_limpo}")
                for row in ws.iter_rows(min_row=2, values_only=False):
                    try:
                        nome_excel = normalize_name(str(row[0].value).strip())
                        if nome_procurado_limpo == nome_excel:
                            row[pago_col_index - 1].value = "X"
                            encontrado = True
                            print(f"Encontrado e marcado como pago: {nome_procurado_limpo}")
                            break
                    except Exception as e:
                        print(f"Erro ao processar linha do Excel: {str(e)}")
                        continue
                if encontrado:
                    resultados.append((nome_procurado, "Encontrado e marcado como pago"))
                else:
                    resultados.append((nome_procurado, "Não encontrado"))
            except Exception as e:
                print(f"Erro ao processar nome do débito automático: {str(e)}")
                continue

        wb.save(filename=file_path_excel)
        wb.close()

        txt_area.delete('1.0', tk.END)
        txt_area.insert(tk.END, "{:<50} {:<30}\n".format("Nomes processados do arquivo RET", "Status"))
        txt_area.insert(tk.END, "="*80 + "\n")
        for nome, status in resultados:
            txt_area.insert(tk.END, "{:<50} {:<30}\n".format(nome, status))
        txt_area.insert(tk.END, "Processamento concluído.\n")
    except Exception as e:
        print(f"Erro ao processar arquivo RET: {str(e)}")

def processar_arquivos(file_path_txt, file_path_excel, aba, txt_area, file_path_debito_automatico=None):
    try:
        with open(file_path_txt, 'r', encoding='utf-8') as file:
            content_txt = file.readlines()
        data_posicao = extrair_data_posicao_csv(file_path_txt)
        nomes_valores_txt = extrair_nomes_e_valores(content_txt)
        nomes = [normalize_name(nome) for nome, _ in nomes_valores_txt]

        taxas = extrair_e_somar_duplas(content_txt)

        nomes_debito_automatico = []
        if file_path_debito_automatico:
            nomes_debito_automatico = extrair_nomes_debito_automatico(file_path_debito_automatico)
            nomes.extend(nomes_debito_automatico)

        nomes_encontrados, _ = buscar_nomes_excel(file_path_excel, nomes, aba, data_posicao, taxas, nomes_debito_automatico)
        txt_area.delete('1.0', tk.END)
        txt_area.insert(tk.END, "{:<50} {:<30}\n".format("Nomes processados do arquivo TXT", "Status"))
        txt_area.insert(tk.END, "="*80 + "\n")
        for nome in nomes:
            status = "X" if nome in nomes_encontrados else "Não encontrado"
            txt_area.insert(tk.END, "{:<50} {:<30}\n".format(nome, status))
        quantidade_nomes = len(nomes_encontrados)
        txt_area.insert(tk.END, f"Quantidade de nomes correspondentes encontrados: {quantidade_nomes}\n")
    except Exception as e:
        print(f"Erro ao processar arquivos TXT e RET: {str(e)}")

def extrair_data_posicao_csv(file_path_txt):
    try:
        with open(file_path_txt, 'r', encoding='utf-8') as file:
            content = file.readlines()
        for line in content:
            if 'POSICAO DO DIA:' in line:
                try:
                    data = line.split(':')[-1].strip()
                    return data
                except IndexError:
                    continue
        return None
    except Exception as e:
        print(f"Erro ao extrair data de posição do arquivo CSV: {str(e)}")
        return None

def extrair_e_somar_duplas(content):
    try:
        taxas = {}
        for i, line in enumerate(content):
            if '/01' in line and any(char.isalpha() for char in line):
                try:
                    partes = line.split()
                    nome = []
                    for parte in partes[3:]:
                        if parte.replace(',', '').replace('.', '').isdigit():
                            break
                        nome.append(parte)
                    nome = ' '.join(nome).strip()
                    taxas[nome] = 0

                    soma = 0
                    for offset in range(1, 3):
                        if i + offset < len(content):
                            next_line = content[i + offset]
                            matches = re.findall(r'(41|07)\s+(\d+,\d+)', next_line)
                            for match in matches:
                                _, numero = match
                                numero = float(numero.replace(',', '.'))
                                soma += numero
                    taxas[nome] = soma
                except Exception as e:
                    print(f"Erro ao extrair e somar duplas: {str(e)}")
                    continue
        return taxas
    except Exception as e:
        print(f"Erro ao extrair e somar duplas: {str(e)}")
        return {}

def extrair_nomes_e_valores(content):
    try:
        nomes = []
        valores = []
        for line in content:
            if '/01' in line and any(char.isalpha() for char in line):
                try:
                    partes = line.split()
                    nome = []
                    for parte in partes[3:]:
                        if parte.replace('.', '').replace(',', '').isdigit():
                            break
                        nome.append(parte)
                    nomes.append(' '.join(nome).strip())

                    next_line_index = content.index(line) + 1
                    if next_line_index < len(content):
                        next_line = content[next_line_index]
                        if 'DINHEIRO' in next_line or 'D ' in next_line:
                            try:
                                if 'C' in next_line:
                                    index_c = next_line.index('C')
                                    valor_titulo = next_line[index_c + 1:].strip().split()[0].replace(',', '.')
                                elif 'D ' in next_line:
                                    index_d = next_line.index('D ')
                                    valor_titulo = next_line[index_d + 1:].strip().split()[0].replace(',', '.')
                                else:
                                    continue
                                valores.append(valor_titulo)
                            except (ValueError, IndexError):
                                valores.append("Valor não encontrado")
                        else:
                            valores.append("Valor não encontrado")
                    else:
                        valores.append("Valor não encontrado")
                except Exception as e:
                    print(f"Erro ao extrair nomes e valores: {str(e)}")
                    continue
        return list(zip(nomes, valores))
    except Exception as e:
        print(f"Erro ao extrair nomes e valores do arquivo CSV: {str(e)}")
        return []



