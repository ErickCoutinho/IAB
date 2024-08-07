import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox, ttk
from PIL import Image, ImageTk, ImageFont
from openpyxl import load_workbook
import threading

# Funções de processamento
from funçoes import processar_arquivos, processar_arquivo_ret

def selecionar_arquivo_txt():
    try:
        file_path = filedialog.askopenfilename(filetypes=[("Arquivos TXT", "*.txt")])
        if file_path:
            txt_path_entry.delete(0, tk.END)
            txt_path_entry.insert(tk.END, file_path)
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao selecionar arquivo TXT: {str(e)}")

def selecionar_arquivo_excel():
    try:
        file_path = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx;*.xls")])
        if file_path:
            excel_path_entry.delete(0, tk.END)
            excel_path_entry.insert(tk.END, file_path)
            try:
                wb = load_workbook(filename=file_path, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
                aba_combo['values'] = sheet_names  # Atualiza o combobox com as abas disponíveis
            except FileNotFoundError:
                messagebox.showerror("Erro", f"O arquivo Excel '{file_path}' não foi encontrado.")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao carregar abas do Excel: {str(e)}")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao selecionar arquivo Excel: {str(e)}")

def selecionar_arquivo_debito():
    try:
        file_path = filedialog.askopenfilename(filetypes=[("Arquivos RET", "*.ret")])
        if file_path:
            debito_path_entry.delete(0, tk.END)
            debito_path_entry.insert(tk.END, file_path)
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao selecionar arquivo de Débito Automático: {str(e)}")

def mostrar_loading():
    try:
        global loading_window
        loading_window = tk.Toplevel(root)
        loading_window.title("Carregando...")

        # Centralizar a janela
        largura = 300
        altura = 150
        largura_tela = loading_window.winfo_screenwidth()
        altura_tela = loading_window.winfo_screenheight()
        x = (largura_tela / 2) - (largura / 2)
        y = (altura_tela / 2) - (altura / 2)
        loading_window.geometry(f"{largura}x{altura}+{int(x)}+{int(y)}")

        # Adicionar imagem de carregamento
        loading_image = Image.open("SELO+IABRS COMPL_BRANCO.png")
        loading_image = loading_image.resize((50, 50), Image.LANCZOS)
        loading_photo = ImageTk.PhotoImage(loading_image)
        loading_label_image = tk.Label(loading_window, image=loading_photo)
        loading_label_image.image = loading_photo  # Manter uma referência para evitar coleta de lixo
        loading_label_image.pack(pady=10)

        loading_label_text = tk.Label(loading_window, text="Processando, por favor aguarde...", font=("Arial", 12))
        loading_label_text.pack()

        progress_bar = ttk.Progressbar(loading_window, mode='indeterminate')
        progress_bar.pack(expand=True, fill=tk.BOTH, side=tk.BOTTOM, padx=20, pady=20)
        progress_bar.start()
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao mostrar a janela de carregamento: {str(e)}")

def esconder_loading():
    try:
        loading_window.destroy()
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao esconder a janela de carregamento: {str(e)}")

def processar():
    try:
        file_path_txt = txt_path_entry.get() if txt_path_entry.get() else None
        file_path_excel = excel_path_entry.get()
        aba = aba_combo.get()

        file_path_debito = debito_path_entry.get() if debito_path_entry.get() else None

        if (file_path_txt or file_path_debito) and file_path_excel and aba:
            threading.Thread(target=executar_processamento,
                            args=(file_path_txt, file_path_excel, aba, file_path_debito)).start()
        else:
            messagebox.showerror("Erro",
                                "Selecione um arquivo TXT ou um arquivo de Débito Automático, um arquivo Excel e uma aba.")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao iniciar o processamento: {str(e)}")

def executar_processamento(file_path_txt, file_path_excel, aba, file_path_debito):
    root.after(0, mostrar_loading)
    try:
        if file_path_txt:
            processar_arquivos(file_path_txt, file_path_excel, aba, txt_area, file_path_debito)
        elif file_path_debito:
            processar_arquivo_ret(file_path_debito, file_path_excel, aba, txt_area)
        else:
            raise ValueError("Nenhum arquivo TXT ou de Débito Automático foi selecionado.")
    except Exception as e:
        messagebox.showerror("Erro", str(e))
    finally:
        root.after(0, esconder_loading)

# Criar a interface gráfica
root = tk.Tk()
root.title("Selecionar Arquivos e Abas de Arquivos Excel")
root.configure(bg="black")

# Carregar a logo
try:
    logo_image = Image.open("SELO+IABRS COMPL_BRANCO.png")
    logo_image = logo_image.resize((200, 150), Image.LANCZOS)
    logo_photo = ImageTk.PhotoImage(logo_image)

    # Fonte DIN (certifique-se de que o arquivo .ttf está no mesmo diretório ou forneça o caminho completo)
    font_path = "DIN-Regular.ttf"
    din_font = ImageFont.truetype(font_path, 12)
except Exception as e:
    messagebox.showerror("Erro", f"Erro ao carregar a logo ou a fonte: {str(e)}")

# Frame para a logo
logo_frame = tk.Frame(root, bg="black")
logo_frame.pack(pady=10)
logo_label = tk.Label(logo_frame, image=logo_photo, bg="black")
logo_label.image = logo_photo  # Keep a reference to avoid garbage collection
logo_label.pack()

# Configurações do frame principal
frame = tk.Frame(root, bg="black")
frame.pack(padx=20, pady=20)

labels = ["Selecione o arquivo Bancário (.txt):", "Selecione o arquivo de Débito Automático (.ret):", "Selecione o arquivo Excel:", "Selecione a aba do arquivo Excel:"]
for i, text in enumerate(labels):
    tk.Label(frame, text=text, fg="white", bg="black", font=("DIN", 12)).grid(row=i, column=0, sticky='w', pady=5)

txt_path_entry = tk.Entry(frame, width=50, font=("DIN", 12))
txt_path_entry.grid(row=0, column=1, padx=10)
btn_selecionar_txt = tk.Button(frame, text="Selecionar", command=selecionar_arquivo_txt, font=("DIN", 12))
btn_selecionar_txt.grid(row=0, column=2, padx=10)

debito_path_entry = tk.Entry(frame, width=50, font=("DIN", 12))
debito_path_entry.grid(row=1, column=1, padx=10)
btn_selecionar_debito = tk.Button(frame, text="Selecionar", command=selecionar_arquivo_debito, font=("DIN", 12))
btn_selecionar_debito.grid(row=1, column=2, padx=10)

excel_path_entry = tk.Entry(frame, width=50, font=("DIN", 12))
excel_path_entry.grid(row=2, column=1, padx=10)
btn_selecionar_excel = tk.Button(frame, text="Selecionar", command=selecionar_arquivo_excel, font=("DIN", 12))
btn_selecionar_excel.grid(row=2, column=2, padx=10)

aba_combo = ttk.Combobox(frame, width=47, state='readonly', font=("DIN", 12))
aba_combo.grid(row=3, column=1, padx=10, columnspan=2)

btn_processar = tk.Button(root, text="Processar Arquivos", command=processar, font=("DIN", 12))
btn_processar.pack(pady=10)

txt_area = scrolledtext.ScrolledText(root, width=100, height=20, font=("DIN", 12), bg="black", fg="white")
txt_area.pack()

root.mainloop()
